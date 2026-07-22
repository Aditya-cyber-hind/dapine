import csv
import json
import os
from datetime import date, datetime

class Connectors:
    def __init__(self):
        self.connections = {}
    
    # ============ DATABASE ============
    
    def read_database(self, connection_string, query, alias):
        try:
            from sqlalchemy import create_engine, text
            engine = create_engine(connection_string)
            with engine.connect() as conn:
                result = conn.execute(text(query))
                rows = [dict(row._mapping) for row in result]
                columns = list(rows[0].keys()) if rows else []
                for row in rows:
                    for k, v in row.items():
                        if isinstance(v, (date, datetime)):
                            row[k] = str(v)
                return rows, columns
        except Exception as e:
            raise Exception(f"Database error: {e}")
    
    def write_database(self, rows, columns, connection_string, table_name):
        try:
            from sqlalchemy import create_engine, MetaData, Table, Column, String
            from sqlalchemy.orm import sessionmaker
            engine = create_engine(connection_string)
            metadata = MetaData()
            cols = [Column(c, String(255)) for c in columns]
            table = Table(table_name, metadata, *cols)
            metadata.create_all(engine)
            Session = sessionmaker(bind=engine)
            session = Session()
            for row in rows:
                session.execute(table.insert().values(**{k: str(v) if v is not None else None for k, v in row.items()}))
            session.commit()
            session.close()
            return True
        except Exception as e:
            raise Exception(f"Database write error: {e}")
    
    # ============ EXCEL ============
    
    def read_excel(self, filepath, sheet_name=None):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filepath, data_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active
            rows_data = []
            headers = None
            for row in ws.iter_rows(values_only=True):
                if headers is None:
                    headers = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(row)]
                else:
                    row_dict = {}
                    for i, val in enumerate(row):
                        if i < len(headers):
                            row_dict[headers[i]] = self._infer_excel_type(val)
                    rows_data.append(row_dict)
            wb.close()
            return rows_data, headers
        except Exception as e:
            raise Exception(f"Excel read error: {e}")
    
    def write_excel(self, rows, columns, filepath, sheet_name="Sheet1"):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font; cell.fill = header_fill
            for row_idx, row in enumerate(rows, 2):
                for col_idx, col_name in enumerate(columns, 1):
                    val = row.get(col_name)
                    if isinstance(val, (date, datetime)): val = str(val)
                    ws.cell(row=row_idx, column=col_idx, value=val)
            for col_idx, col_name in enumerate(columns, 1):
                max_length = len(str(col_name))
                for row in rows:
                    max_length = max(max_length, len(str(row.get(col_name, ""))))
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
            wb.save(filepath); wb.close()
            return True
        except Exception as e:
            raise Exception(f"Excel write error: {e}")
    
    # ============ PARQUET ============
    
    def read_parquet(self, filepath):
        try:
            import pandas as pd
            df = pd.read_parquet(filepath)
            rows = df.to_dict('records')
            columns = list(df.columns)
            for row in rows:
                for k, v in row.items():
                    if isinstance(v, (date, datetime)): row[k] = str(v)
                    elif hasattr(v, 'item'): row[k] = v.item()
            return rows, columns
        except Exception as e:
            raise Exception(f"Parquet read error: {e}")
    
    def write_parquet(self, rows, columns, filepath):
        try:
            import pandas as pd
            df = pd.DataFrame(rows, columns=columns)
            df.to_parquet(filepath, index=False)
            return True
        except Exception as e:
            raise Exception(f"Parquet write error: {e}")
    
    # ============ GOOGLE SHEETS ============
    
    def read_google_sheet(self, sheet_url, sheet_name=None):
        try:
            import gspread
            from oauth2client.service_account import ServiceAccountCredentials
            scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
            creds_file = os.environ.get('GOOGLE_CREDS', 'google_creds.json')
            if not os.path.exists(creds_file):
                return [], []
            creds = ServiceAccountCredentials.from_json_keyfile_name(creds_file, scope)
            client = gspread.authorize(creds)
            sheet = client.open_by_url(sheet_url)
            worksheet = sheet.worksheet(sheet_name) if sheet_name else sheet.sheet1
            data = worksheet.get_all_records()
            if data:
                columns = list(data[0].keys())
                return data, columns
            return [], []
        except Exception as e:
            print(f"Google Sheets read error: {e}")
            return [], []
    
    def _infer_excel_type(self, val):
        if val is None: return None
        if isinstance(val, (date, datetime)): return str(val)
        if isinstance(val, (int, float)): return val
        return str(val)