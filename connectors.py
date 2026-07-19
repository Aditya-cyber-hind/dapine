import csv
import json
import os
from datetime import date, datetime

class Connectors:
    def __init__(self):
        self.connections = {}
    
    # ============ DATABASE ============
    
    def read_database(self, connection_string, query, alias):
        """Read from any SQL database."""
        try:
            from sqlalchemy import create_engine, text
            
            engine = create_engine(connection_string)
            with engine.connect() as conn:
                result = conn.execute(text(query))
                rows = [dict(row._mapping) for row in result]
                columns = list(rows[0].keys()) if rows else []
                
                # Convert types
                for row in rows:
                    for k, v in row.items():
                        if isinstance(v, (date, datetime)):
                            row[k] = str(v)
                
                return rows, columns
        except Exception as e:
            raise Exception(f"Database error: {e}")
    
    def write_database(self, rows, columns, connection_string, table_name):
        """Write to any SQL database."""
        try:
            from sqlalchemy import create_engine, Table, Column, MetaData, String
            from sqlalchemy.orm import sessionmaker
            
            engine = create_engine(connection_string)
            metadata = MetaData()
            
            # Create table dynamically
            cols = [Column(c, String(255)) for c in columns]
            table = Table(table_name, metadata, *cols)
            metadata.create_all(engine)
            
            Session = sessionmaker(bind=engine)
            session = Session()
            
            for row in rows:
                session.execute(table.insert().values(**{k: str(v) if v is not None else None for k, v in row.items()}))
            
            session.commit()
            session.close()
            return True
        except Exception as e:
            raise Exception(f"Database write error: {e}")
    
    # ============ EXCEL ============
    
    def read_excel(self, filepath, sheet_name=None):
        """Read Excel file."""
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(filepath, data_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active
            
            rows_data = []
            headers = None
            
            for row in ws.iter_rows(values_only=True):
                if headers is None:
                    headers = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(row)]
                else:
                    row_dict = {}
                    for i, val in enumerate(row):
                        if i < len(headers):
                            row_dict[headers[i]] = self._infer_excel_type(val)
                    rows_data.append(row_dict)
            
            wb.close()
            return rows_data, headers
        except Exception as e:
            raise Exception(f"Excel read error: {e}")
    
    def write_excel(self, rows, columns, filepath, sheet_name="Sheet1"):
        """Write Excel file."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Header style
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
            header_align = Alignment(horizontal="center")
            
            # Write headers
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            
            # Write data
            for row_idx, row in enumerate(rows, 2):
                for col_idx, col_name in enumerate(columns, 1):
                    val = row.get(col_name)
                    if isinstance(val, (date, datetime)):
                        val = str(val)
                    ws.cell(row=row_idx, column=col_idx, value=val)
            
            # Auto-adjust column widths
            for col_idx, col_name in enumerate(columns, 1):
                max_length = len(str(col_name))
                for row in rows:
                    val = str(row.get(col_name, ""))
                    max_length = max(max_length, len(val))
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
            
            wb.save(filepath)
            wb.close()
            return True
        except Exception as e:
            raise Exception(f"Excel write error: {e}")
    
    def _infer_excel_type(self, val):
        if val is None:
            return None
        if isinstance(val, (date, datetime)):
            return str(val)
        if isinstance(val, (int, float)):
            return val
        return str(val)