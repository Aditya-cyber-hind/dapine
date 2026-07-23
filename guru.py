import sys
import os
import json
import urllib.request
import csv
import webbrowser

GREEN = '\033[92m'
CYAN = '\033[96m'
YELLOW = '\033[93m'
RED = '\033[91m'
BOLD = '\033[1m'
RESET = '\033[0m'

def print_banner():
    print(f"""
{CYAN}{BOLD}
  ╔══════════════════════════════════════════╗
  ║        🧘 GURU - Dapine AI Assistant     ║
  ║    "Talk to your data. I'll do the rest." ║
  ╚══════════════════════════════════════════╝
{RESET}
  {GREEN}/load file{RESET}   - Load a CSV or Excel file
  {GREEN}/ask question{RESET} - Ask anything about your data
  {GREEN}/show{RESET}        - Show current data
  {GREEN}/chart{RESET}       - Create a chart
  {GREEN}/help{RESET}        - Show this message
  {GREEN}/exit{RESET}        - Goodbye
""")

class Guru:
    def __init__(self):
        self.data = None
        self.columns = None
        self.filename = None
        self.groq_client = None
        self._init_groq()
    
    def _init_groq(self):
        api_key = os.environ.get('GROQ_API_KEY', '')
        if api_key:
            try:
                from groq import Groq
                self.groq_client = Groq(api_key=api_key)
                print(f"{GREEN}🧘 Guru: AI mode activated!{RESET}")
            except:
                pass
    
    def load_file(self, filepath):
        if not os.path.exists(filepath):
            print(f"{RED}❌ File not found: {filepath}{RESET}")
            return
        
        self.filename = filepath
        
        if filepath.endswith('.csv'):
            with open(filepath, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                self.data = []
                for row in reader:
                    typed_row = {}
                    for k, v in row.items():
                        v = v.strip()
                        try: typed_row[k] = int(v)
                        except:
                            try: typed_row[k] = float(v)
                            except: typed_row[k] = v
                    self.data.append(typed_row)
                self.columns = list(self.data[0].keys()) if self.data else []
        
        elif filepath.endswith('.xlsx'):
            from openpyxl import load_workbook
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            self.columns = [str(c) for c in rows[0]]
            self.data = []
            for row in rows[1:]:
                typed_row = {}
                for i, val in enumerate(row):
                    if i < len(self.columns):
                        if isinstance(val, (int, float)):
                            typed_row[self.columns[i]] = val
                        else:
                            typed_row[self.columns[i]] = str(val) if val else None
                self.data.append(typed_row)
            wb.close()
        else:
            print(f"{RED}❌ Unsupported format. Use .csv or .xlsx{RESET}")
            return
        
        print(f"{GREEN}✅ Loaded {len(self.data)} rows from {filepath}{RESET}")
        print(f"{CYAN}📊 Columns ({len(self.columns)}): {', '.join(self.columns)}{RESET}")
        self._show_sample()
    
    def _show_sample(self):
        if self.data:
            print(f"{CYAN}📋 Sample (first 3 rows):{RESET}")
            for row in self.data[:3]:
                # Show only first 5 columns
                keys = list(row.keys())[:5]
                sample = {k: row[k] for k in keys}
                print(f"  {sample}")
            print()
    
    def ask(self, question):
        if not self.data:
            print(f"{YELLOW}⚠️  No data loaded. Use /load first.{RESET}")
            return
        
        q = question.lower()
        
        # Show data
        if any(w in q for w in ['show', 'display', 'print', 'see', 'view']):
            if 'all' in q or 'everything' in q:
                self._print_table(self.data[:20])
            elif 'first' in q:
                n = self._extract_number(q) or 5
                self._print_table(self.data[:n])
            elif 'last' in q:
                n = self._extract_number(q) or 5
                self._print_table(self.data[-n:])
            elif 'sample' in q:
                self._show_sample()
            else:
                self._print_table(self.data[:10])
            return
        
        # Row/column count
        if any(w in q for w in ['how many', 'count', 'number of']):
            if 'row' in q or 'record' in q or 'entry' in q:
                print(f"{GREEN}📊 {len(self.data)} rows{RESET}")
            elif 'column' in q or 'field' in q:
                print(f"{GREEN}📊 {len(self.columns)} columns: {', '.join(self.columns)}{RESET}")
            return
        
        # What columns
        if any(w in q for w in ['what column', 'what field', 'columns', 'fields']):
            print(f"{GREEN}📊 Columns: {', '.join(self.columns)}{RESET}")
            return
        
        # Stats
        if any(w in q for w in ['stat', 'average', 'mean', 'min', 'max', 'sum', 'total', 'median']):
            self._smart_stats(q)
            return
        
        # Filter
        if any(w in q for w in ['filter', 'find', 'where', 'which', 'show me']):
            if self._smart_filter(q):
                return
        
        # Sort / top / bottom
        if any(w in q for w in ['top', 'bottom', 'highest', 'lowest', 'sort', 'order', 'rank']):
            self._smart_sort(q)
            return
        
        # Group
        if any(w in q for w in ['group', 'per', 'by', 'breakdown', 'aggregate']):
            self._smart_group(q)
            return
        
        # Chart
        if any(w in q for w in ['chart', 'graph', 'plot', 'visualize', 'bar', 'pie', 'line']):
            self._smart_chart(q)
            return
        
        # AI fallback
        if self.groq_client:
            self._ask_groq(question)
        else:
            print(f"{YELLOW}🤔 Guru: I don't understand. Try:{RESET}")
            print(f"  - 'show me the data'")
            print(f"  - 'what's the average price'")
            print(f"  - 'top 5 by sales'")
            print(f"  - 'show me where age > 25'")
            print(f"  - 'bar chart of revenue by region'")
            print(f"{CYAN}  Set GROQ_API_KEY for AI-powered answers!{RESET}")
    
    def _extract_number(self, text):
        import re
        match = re.search(r'\b(\d+)\b', text)
        return int(match.group(1)) if match else None
    
    def _get_numeric_columns(self):
        num_cols = []
        for col in self.columns:
            vals = [r.get(col) for r in self.data[:10] if r.get(col) is not None]
            if vals and all(isinstance(v, (int, float)) for v in vals):
                num_cols.append(col)
        return num_cols
    
    def _print_table(self, rows, max_rows=20):
        if not rows:
            print(f"{YELLOW}No data to show.{RESET}")
            return
        print(f"\n{GREEN}📊 Showing {len(rows)} rows:{RESET}")
        print(f"{'─'*60}")
        for row in rows[:max_rows]:
            keys = list(row.keys())[:5]
            small = {k: row[k] for k in keys}
            if len(row) > 5:
                print(f"  {small} ...")
            else:
                print(f"  {small}")
        if len(rows) > max_rows:
            print(f"  ... and {len(rows) - max_rows} more")
        print(f"{'─'*60}\n")
    
    def _smart_filter(self, question):
        q = question.lower()
        
        for col in self.columns:
            if col.lower() in q:
                import re
                
                # Greater than
                match = re.search(rf'{col}\s*(?:is\s*)?(?:greater|more|over|above|>|higher)\s*(?:than\s*)?(\d+\.?\d*)', q)
                if match:
                    val = float(match.group(1))
                    filtered = [r for r in self.data if r.get(col) is not None and float(r.get(col, 0)) > val]
                    print(f"{GREEN}🔍 {col} > {val}: {len(filtered)} rows found{RESET}")
                    self._print_table(filtered[:10])
                    return True
                
                # Less than
                match = re.search(rf'{col}\s*(?:is\s*)?(?:less|under|below|<|lower)\s*(?:than\s*)?(\d+\.?\d*)', q)
                if match:
                    val = float(match.group(1))
                    filtered = [r for r in self.data if r.get(col) is not None and float(r.get(col, 0)) < val]
                    print(f"{GREEN}🔍 {col} < {val}: {len(filtered)} rows found{RESET}")
                    self._print_table(filtered[:10])
                    return True
                
                # Equals
                match = re.search(rf'{col}\s*(?:is\s*|=\s*|equals?\s*)["\']?(\w+)["\']?', q)
                if match:
                    val = match.group(1)
                    filtered = [r for r in self.data if str(r.get(col, '')).lower() == val.lower()]
                    print(f"{GREEN}🔍 {col} = '{val}': {len(filtered)} rows found{RESET}")
                    self._print_table(filtered[:10])
                    return True
        
        return False
    
    def _smart_stats(self, question):
        q = question.lower()
        
        for col in self.columns:
            if col.lower() in q:
                vals = [r.get(col) for r in self.data if r.get(col) is not None]
                try:
                    vals = [float(v) for v in vals]
                except:
                    print(f"{YELLOW}📊 {col} is not numeric{RESET}")
                    return
                
                if vals:
                    import statistics
                    print(f"\n{GREEN}📈 Statistics for {col} ({len(vals)} values):{RESET}")
                    print(f"  {'─'*40}")
                    print(f"  Sum:      {sum(vals):>15,.2f}")
                    print(f"  Average:  {sum(vals)/len(vals):>15,.2f}")
                    print(f"  Min:      {min(vals):>15,.2f}")
                    print(f"  Max:      {max(vals):>15,.2f}")
                    if len(vals) > 1:
                        print(f"  Median:   {statistics.median(vals):>15,.2f}")
                        print(f"  Std Dev:  {statistics.stdev(vals):>15,.2f}")
                    print(f"  {'─'*40}\n")
                return
        
        # No specific column found - show stats for first numeric column
        num_cols = self._get_numeric_columns()
        if num_cols:
            col = num_cols[0]
            vals = [float(r.get(col, 0)) for r in self.data if r.get(col) is not None]
            import statistics
            print(f"\n{GREEN}📈 Statistics for {col} ({len(vals)} values):{RESET}")
            print(f"  Average: {sum(vals)/len(vals):,.2f}")
            print(f"  Min: {min(vals):,.2f}  |  Max: {max(vals):,.2f}\n")
    
    def _smart_sort(self, question):
        q = question.lower()
        
        for col in self.columns:
            if col.lower() in q:
                n = self._extract_number(q) or 5
                reverse = any(w in q for w in ['top', 'highest', 'most', 'desc', 'descending'])
                sorted_data = sorted(self.data, key=lambda r: float(r.get(col, 0) or 0), reverse=reverse)
                
                print(f"{GREEN}📊 {'Top' if reverse else 'Bottom'} {min(n, len(sorted_data))} by {col}:{RESET}")
                self._print_table(sorted_data[:n])
                return
        
        print(f"{YELLOW}🤔 Which column? Try: 'top 5 by price'{RESET}")
    
    def _smart_group(self, question):
        q = question.lower()
        
        for col in self.columns:
            if col.lower() in q:
                groups = {}
                for row in self.data:
                    key = str(row.get(col, 'Unknown'))
                    if key not in groups: groups[key] = []
                    groups[key].append(row)
                
                print(f"\n{GREEN}📊 Grouped by {col}:{RESET}")
                print(f"  {'─'*30}")
                for key, rows in sorted(groups.items(), key=lambda x: len(x[1]), reverse=True)[:10]:
                    print(f"  {str(key):<20} {len(rows):>5} rows")
                print(f"  {'─'*30}\n")
                return
        
        print(f"{YELLOW}🤔 Try: 'group by city' or 'breakdown by category'{RESET}")
    
    def _smart_chart(self, question):
        q = question.lower()
        
        chart_type = 'bar'
        if 'pie' in q: chart_type = 'pie'
        if 'line' in q: chart_type = 'line'
        if 'scatter' in q: chart_type = 'scatter'
        
        label_col = None
        value_col = None
        
        for col in self.columns:
            if col.lower() in q:
                if not label_col: label_col = col
                elif not value_col: value_col = col
        
        if not label_col:
            label_col = self.columns[0]
        if not value_col:
            num_cols = self._get_numeric_columns()
            value_col = num_cols[0] if num_cols else self.columns[-1]
        
        if label_col and value_col:
            # Group by label, average the value
            groups = {}
            for row in self.data:
                key = str(row.get(label_col, ''))
                val = row.get(value_col)
                if val is not None and isinstance(val, (int, float)):
                    if key not in groups: groups[key] = []
                    groups[key].append(val)
            
            labels = list(groups.keys())[:15]
            values = [sum(v)/len(v) for v in groups.values()][:15]
            
            output_file = f"guru_{chart_type}_{value_col}_by_{label_col}.html"
            self._generate_chart(labels, values, label_col, value_col, chart_type, output_file)
        else:
            print(f"{YELLOW}🤔 Try: 'bar chart of sales by region'{RESET}")
    
    def _generate_chart(self, labels, values, label_col, value_col, chart_type, output_file):
        colors = [
            "rgba(54,162,235,0.8)", "rgba(255,99,132,0.8)", "rgba(75,192,192,0.8)",
            "rgba(255,206,86,0.8)", "rgba(153,102,255,0.8)", "rgba(255,159,64,0.8)"
        ]
        bg_colors = [colors[i % len(colors)] for i in range(len(labels))]
        
        html = f"""<!DOCTYPE html>
<html><head><title>Guru: {value_col} by {label_col}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>body{{background:#1a1a2e;display:flex;justify-content:center;align-items:center;min-height:100vh;margin:0;font-family:sans-serif}}
.container{{background:#16213e;padding:30px;border-radius:15px;width:80%;max-width:900px}}
h1{{color:#e94560;text-align:center;margin-bottom:10px}}
.subtitle{{color:#8b949e;text-align:center;margin-bottom:30px}}
canvas{{max-height:500px}}</style></head>
<body><div class="container">
<h1>🧘 Guru Chart</h1>
<p class="subtitle">{value_col} by {label_col} ({chart_type} chart)</p>
<canvas id="c"></canvas></div>
<script>
new Chart(document.getElementById('c'),{{type:'{chart_type}',
data:{{labels:{json.dumps(labels)},datasets:[{{label:'{value_col}',data:{json.dumps(values)},
backgroundColor:{json.dumps(bg_colors)},borderColor:'#fff',borderWidth:2,borderRadius:8}}]}},
options:{{responsive:true,plugins:{{legend:{{labels:{{color:'#c9d1d9',font:{{size:14}}}}}}}},
scales:{{y:{{ticks:{{color:'#8b949e'}},grid:{{color:'#21262d'}}}},
x:{{ticks:{{color:'#8b949e'}},grid:{{color:'#21262d'}}}}}}}}}});
</script></body></html>"""
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(html)
        
        webbrowser.open(f"file:///{os.path.abspath(output_file)}")
        print(f"{GREEN}✅ Chart saved to {output_file} and opened in browser!{RESET}")
    
    def _ask_groq(self, question):
        try:
            context = f"Data has {len(self.data)} rows. Columns: {', '.join(self.columns)}. "
            context += f"Sample: {json.dumps(self.data[:3], default=str)}"
            context += f"\nNumeric columns: {', '.join(self._get_numeric_columns())}"
            
            response = self.groq_client.chat.completions.create(
                model="llama-3.1-8b-instant",
                messages=[
                    {"role": "system", "content": "You are Guru, a helpful data assistant. Answer questions about the data clearly. Be concise. Suggest what the user should ask next. Use emojis."},
                    {"role": "user", "content": f"Data context:\n{context}\n\nUser question: {question}"}
                ],
                temperature=0.7,
                max_tokens=250
            )
            
            answer = response.choices[0].message.content
            print(f"\n{GREEN}🧘 Guru says:{RESET}")
            for line in answer.split('\n'):
                print(f"  {line}")
            print()
        except Exception as e:
            print(f"{YELLOW}🤔 Guru: I had trouble connecting to AI. Try rephrasing!{RESET}")
            print(f"  - 'show me the data'")
            print(f"  - 'what's the average price'")
            print(f"  - 'top 5 by sales'")
    
    def start(self):
        print_banner()
        
        while True:
            try:
                user_input = input(f"{BOLD}{GREEN}guru> {RESET}").strip()
                
                if not user_input:
                    continue
                
                if user_input.startswith('/load '):
                    self.load_file(user_input[6:].strip())
                elif user_input.startswith('/ask '):
                    self.ask(user_input[5:].strip())
                elif user_input == '/show':
                    self._print_table(self.data[:20] if self.data else [])
                elif user_input == '/chart':
                    print(f"{YELLOW}Usage: Just type 'bar chart of sales by region'{RESET}")
                elif user_input == '/help':
                    print_banner()
                elif user_input in ('/exit', '/quit', 'exit', 'quit'):
                    print(f"{CYAN}🧘 Guru: Come back with more data! Namaste. 🙏{RESET}")
                    break
                else:
                    self.ask(user_input)
                    
            except KeyboardInterrupt:
                print(f"\n{CYAN}🧘 Guru: Come back with more data! Namaste. 🙏{RESET}")
                break
            except Exception as e:
                print(f"{RED}❌ {e}{RESET}")

if __name__ == "__main__":
    guru = Guru()
    guru.start()